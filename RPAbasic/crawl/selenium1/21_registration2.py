print()
# 사업자등록상태 조회 자동화 + 엑셀 작업
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

from bs4 import BeautifulSoup

from openpyxl import load_workbook


browser = webdriver.Chrome()
browser.get(
    "https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/ab/a/a/UTEABAAA13.xml"
)
browser.maximize_window()

# 엑셀 작업
wb = load_workbook("./RPAbasic/crawl/download/business_number.xlsx")
ws = wb.active

# 엑셀 전체 내용 출력(행,열 == x,y)
# 잘 읽어오는지 확인
for x in range(2, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(x, y).value, end=" ")
    print()

# 엑셀 부분 출력
for x in range(2, ws.max_row + 1):
    # 사업자 등록번호 읽어 오기
    bsn = ws.cell(x, 1).value
    print("===사업자등록번호===", bsn)
    browser.find_element(By.ID, "bsno").send_keys(bsn)

    # 조회하기 클릭
    browser.find_element(By.ID, "trigger5").click()
    time.sleep(2)

    # 상태 화면 출력
    # tbody = browser.find_element(By.XPATH, '//*[@id="grid2_body_tbody"]')
    # print(tbody.text) #767-82-00017 부가가치세 일반과세자 입니다. 2022-06-23

    # bs4 사용
    soup = BeautifulSoup(browser.page_source, "lxml")
    tds = soup.select("#grid2_body_tbody > tr > td")

    # 하나의 리스트로 생성
    list1 = []

    # 조회된 정보를 하나의 리스트 구조로 만들기
    for td in tds:
        # print(td.get_text())
        list1.append(td.get_text())

    # 엑셀 저장
    ws.append(list1)

    time.sleep(5)

    del soup

# for문 종료 후 기존 내용 삭제
# 2번 행부터 10개 삭제
ws.delete_rows(2, 10)

# 저장
wb.save("./RPAbasic/crawl/download/business_number.xlsx")


time.sleep(3)
browser.quit()


print()
