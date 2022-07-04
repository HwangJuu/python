print()
# load_workbook : 불러올때
# Workbook : 생성할 때
from openpyxl import load_workbook, Workbook

# 워크북 가져오기
wb = load_workbook("./RPAbasic/excel/sample.xlsx")
ws = wb.active

# 셀 데이터 출력
# 내용의 범위를 알 때 사용
print("시트 범위를 알 때")
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(x, y).value, end=" ")
    print()

print()
# max_row / max_column : 시트(데이터)가 가지고 있는 행, 열 개수를 가져오기
print("시트 행, 열 개수를 모를 때")
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(x, y).value, end=" ")
    print()


print()
