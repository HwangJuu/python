print()
# 실습
# 1번 ~ 9번 학생들 성적 관리

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 방법 1)
# ws.append(["출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총합"])
# ws.append([1, 10, 8, 5, 14, 26, 12])
# ws.append([2, 7, 3, 7, 15, 24, 18])
# ws.append([3, 9, 5, 8, 8, 12, 4])
# ws.append([4, 7, 8, 7, 17, 21, 18])
# ws.append([5, 7, 8, 7, 16, 25, 15])
# ws.append([6, 3, 5, 8, 8, 17, 0])
# ws.append([7, 4, 9, 10, 16, 27, 18])
# ws.append([8, 6, 6, 6, 15, 19, 17])
# ws.append([9, 10, 10, 9, 19, 30, 19])
# ws.append([10, 9, 8, 8, 20, 25, 20])


# 방법 2)
# 타이틀
ws.append(["출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총합"])
# 2차원
scores = [
    [1, 10, 8, 5, 14, 26, 12],
    [2, 7, 3, 7, 15, 24, 18],
    [3, 9, 5, 8, 8, 12, 4],
    [4, 7, 8, 7, 17, 21, 18],
    [5, 7, 8, 7, 16, 25, 15],
    [6, 3, 5, 8, 8, 17, 0],
    [7, 4, 9, 10, 16, 27, 18],
    [8, 6, 6, 6, 15, 19, 17],
    [9, 10, 10, 9, 19, 30, 19],
    [10, 9, 8, 8, 20, 25, 20],
]

for score in scores:
    ws.append(score)

# 퀴즈2 점수 10점으로 변경
for idx, cell in enumerate(ws["D"]):
    if idx == 0:  # 제목 행인 경우 skip
        continue
    cell.value = 10  # 점수 value 값만 10으로 변경

# 총점, 성적 셀 추가
ws["H1"].value = "총점"
ws["I1"].value = "성적"

# 성적 계산
for idx, score in enumerate(scores, start=2):
    # format을 이용해서 값을 지정.
    ws.cell(row=idx, column=8).value = "=sum(B{0}:G{0})".format(idx)
    # 성적
    total = sum(score[1:]) - score[3] + 10
    grade = None
    if total >= 90:
        grade = "A"
    elif total >= 80:
        grade = "B"
    elif total >= 70:
        grade = "C"
    else:
        grade = "D"

    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade

wb.save("./RPAbasic/excel/scores.xlsx")

print()
