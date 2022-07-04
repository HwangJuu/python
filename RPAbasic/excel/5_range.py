print()
from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active

# 한 줄 씩 데이터 삽입 - 리스트 이용
ws.append(["번호", "영어", "수학"])  # 가장 첫줄 타이틀 입력

# 임의로 영어,수학 점수 입력
for i in range(1, 11):
    ws.append([i, random.randint(0, 100), random.randint(0, 100)])


# 2차원
# rows = [
#     ["이름", "생년월일"],
#     ["홍길동", "801020"],
#     ["성춘향", "811020"],
#     ["김지원", "860920"],
#     ["남주혁", "880705"],
# ]

# for row in rows:
#     ws.append(row)


# 첫번째 행 가져오기
# 방법 1)
print("방법 1)", ws["A1"].value, ws["B1"].value, ws["C1"].value)

print()
# 방법 2)
print("방법 2)")
for i in range(1, 2):
    for j in range(1, 4):
        print(ws.cell(i, j).value, end=" ")
print()


print()
# 방법 3)
print("방법 3)")
for j in range(1, 4):
    print(ws.cell(1, j).value, end=" ")
print()


print()
# 방법 4)
first_row = ws[1]
print("방법 4)")
for cell in first_row:
    print(cell.value, end=" ")


print()
# 2~6번 행 가져오기
row_range = ws[2:6]  # 마지막 번호 6 포함
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()


print()
# 3행부터 마지막까지 가져오기
row_range = ws[3 : ws.max_row]  # 마지막 번호 6 포함
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()


print()
# iter_rows() / iter_cols(): 전체 rows, cols 함수
# 하나씩 꺼내 올수있다.
for row in ws.iter_rows():
    print(row[2].value)

print()

for col in ws.iter_cols():
    print(col[2].value)

wb.save("./RPAbasic/excel/range.xlsx")

print()
