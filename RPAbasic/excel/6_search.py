print()
from openpyxl import load_workbook

wb = load_workbook("./RPAbasic/excel/range.xlsx")
ws = wb.active

# 제목 행 제거 : min_row
# 타이틀 지우기. 1번 행은 타이틀, min_row= 어디부터 읽기. 2행부터 읽어오기
for row in ws.iter_rows(min_row=2):
    print(row[0].value, row[1].value, row[2].value)


print()
# 영어 점수가 80 이상인 것만 출력
# for row in ws.iter_rows(min_row=2):
#     if row[1].value > 80:
#         print(row[0].value, row[1].value, row[2].value)

# 셀 안에 '영어'라는 문자가 입력된 셀이 있는 경우 '컴퓨터' 변경
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("./RPAbasic/excel/range.xlsx")

print()
