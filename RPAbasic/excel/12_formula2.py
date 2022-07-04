print()

# 함수 삽입
from openpyxl import load_workbook
from datetime import datetime

# 수식이 들어 있는 경우 수식 그대로 읽어옴
# wb = load_workbook("./RPAbasic/excel/formula.xlsx")
wb = load_workbook("./RPAbasic/excel/formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

print()
